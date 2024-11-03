#!/usr/bin/env python3.8
from flask import Flask, request, send_file, render_template, jsonify, Response, session, make_response
import os
import copy
import traceback
import openpyxl
import xlrd
import xlwt
from xlwt import easyxf
from xlutils.copy import copy as xl_copy
import logging
import io
import threading
import time
from threading import Event
import uuid

app = Flask(__name__)
app.secret_key = 'supersecretkey'  # Replace with a real secret key
UPLOAD_FOLDER = 'uploads'

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

# Set up custom logger for processing
log_streams = {}  # Dictionary to hold log streams for each session
processing_loggers = {}  # Dictionary to hold loggers for each session
log_seek_location = {}  # Dictionary to hold seek locations for each session

# Event to stop processing
stop_events = {}  # Dictionary to hold stop events for each session

def get_session_id():
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    return session['session_id']

class DiscardAfterReadIO(io.StringIO):
    lock = {}

    def __init__(self, initial_value='', newline='\n', session_id=None):
        super().__init__(initial_value, newline)
        session_id = get_session_id()
        if session_id not in self.lock:
            self.lock[session_id] = threading.Lock()
        self._lock = self.lock[session_id]

    def read(self, size=-1):
        with self._lock:
            self.seek(0)
            result = super().read(size)
            self.seek(0)
            self.truncate(0)
            return result

    def readline(self, size=-1):
        with self._lock:
            self.seek(0)
            result = super().readline(size)
            self.seek(0)
            self.truncate(0)
            return result

    def readlines(self, hint=-1):
        with self._lock:
            self.seek(0)
            result = super().readlines(hint)
            self.seek(0)
            self.truncate(0)
            return result

    def write(self, s):
        with self._lock:
            return super().write(s)

    def getvalue(self):
        with self._lock:
            result = super().getvalue()
            self.seek(0)
            self.truncate(0)
            return result

def get_logger(session_id):
    if session_id not in log_streams:
        log_streams[session_id] = DiscardAfterReadIO()
        processing_logger = logging.getLogger(session_id)
        processing_logger.setLevel(logging.INFO)
        log_handler = logging.StreamHandler(log_streams[session_id])
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        log_handler.setFormatter(formatter)
        processing_logger.addHandler(log_handler)
        processing_loggers[session_id] = processing_logger
    return processing_loggers[session_id]

def get_stop_event(session_id):
    if session_id not in stop_events:
        stop_events[session_id] = Event()
    return stop_events[session_id]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_excel(file_path, processing_logger, stop_event):
    processing_logger.info(f"Reading file: {file_path}")
    if stop_event.is_set():
        processing_logger.warning("Processing stopped by user during file read.")
        return None, None, None
    try:
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            return data, wb, sheet
        else:
            book = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = book.sheet_by_index(0)
            data = [sheet.row_values(rowx) for rowx in range(sheet.nrows)]
            return data, book, sheet
    except Exception as e:
        processing_logger.error(f"Error reading file: {e}")
        return None, None, None

def field_verify(fields_dict, data_line):
    for location, field_name in fields_dict.items():
        if field_name not in data_line[location]:
            raise ValueError(f"Field '{field_name}' not found in {data_line}.")

summary_data_field_dic = {
    0: "医疗机构分类",
    1: "医院总数量",
    2: "参加国家卫生健康委员会临床检验中心室间质量评价合格医院数量",
    3: "通过国家室间质评平均合格项目数量",
    4: "参加辽宁省临床检验中心室间质量评价合格医院数量",
    5: "通过辽宁省室间质评平均合格项目数量",
    6: "参加辽宁省医学影像质控中心影像质控认证评价合格医院数量",
    7: "要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可的医院数量",
    8: "认可其他医疗机构标有互认标识的医学影像检查资料和医学检验结果",
    9: "联盟医院间通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频次合计",
    10: "实施检查检验结果互认为患者节约医疗费用",
}

summary_data_field_data_type = {
    summary_data_field_dic[0]: 0,
    summary_data_field_dic[1]: 0,
    summary_data_field_dic[2]: 0,
    summary_data_field_dic[3]: 0,
    summary_data_field_dic[4]: 0,
    summary_data_field_dic[5]: 0,
    summary_data_field_dic[6]: 0,
    summary_data_field_dic[7]: 0,
    summary_data_field_dic[8]: 0,
    summary_data_field_dic[9]: 0,
    summary_data_field_dic[10]: 0.0,
}

large_data_field_dic = {
    0: "医疗机构名称",
    1: "医疗机构分类",
    2: "是否参加国家卫生健康委员会临床检验中心室间质量评价并合格",
    3: "通过国家室间质评医学检验结果合格项目数量",
    4: "是否参加辽宁省临床检验中心室间质量评价并合格",
    5: "通过辽宁省室间质评医学检验结果合格项目数量",
    6: "是否参加辽宁省医学影像质控中心影像质控认证评价并合格",
    7: "要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可",
    8: "DR互认项目数",
    9: "DR节约检查费用",
    10: "MR互认项目数",
    11: "MR节约检查费用",
    12: "CT互认项目数",
    13: "CT节约检查费用",
    14: "临床检验互认项目数",
    15: "临床检验节约检查费用",
    16: "通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频",
    17: "总项目",
    18: "总费用",
}


def validate_summary_data(file_name, summary_data, processing_logger):
    processing_logger.info(f"Validating data for: {file_name}")
    fields_dict = summary_data_field_dic
    data_line = summary_data[3]
    field_verify(fields_dict, data_line)
    if "三级甲等医院" not in summary_data[4][0]:
        raise ValueError("三级甲等医院 is not in row 5, column 1")
    if "三级公立医院" not in summary_data[5][0]:
        raise ValueError("三级公立医院 is not in row 6, column 1")
    if "三级民营医院" not in summary_data[6][0]:
        raise ValueError("三级民营医院 is not in row 7, column 1")
    if "二级公立医院" not in summary_data[7][0]:
        raise ValueError("二级公立医院 is not in row 8, column 1")
    if "二级民营医院" not in summary_data[8][0]:
        raise ValueError("二级民营医院 is not in row 9, column 1")

    processing_logger.info(f"Data validation successful for {file_name}.")


def validate_large_data(file_name, large_data, processing_logger):
    processing_logger.info(f"Validating data for: {file_name}")
    fields_dict = large_data_field_dic
    data_line = large_data[4]
    field_verify(fields_dict, data_line)
    processing_logger.info(f"Data validation successful for {file_name}.")


def get_one_row_data(rows):
    start_yield = False
    for row in rows:
        if not isinstance(row[0], str):
            continue
        if "总计" in row[0]:
            break
    else:
        raise ValueError("No '总计' row found in large Excel data.")

    found = False
    for row in rows:
        if not isinstance(row[0], str):
            continue
        if "总计" in row[0]:
            found = True
        if found and "注" in row[0]:
            break
    else:
        raise ValueError("No '注' row found in large Excel data.")

    for row in rows:
        if not isinstance(row[0], str):
            continue
        if "总计" in row[0]:
            start_yield = True
            continue
        if "注" in row[0]:
            return
        if start_yield:
            yield_dict = dict()
            for idx, value in enumerate(row):
                if idx >= len(large_data_field_dic):
                    break
                if not value:
                    value = 0
                if value == r'/':
                    value = 0
                dict_key = large_data_field_dic[idx]
                yield_dict[dict_key] = value
            yield yield_dict


def summarize_large_data(large_data, processing_logger, stop_event):
    processing_logger.info("Summarizing large Excel data")
    if stop_event.is_set():
        processing_logger.warning("Processing stopped by user during summarization.")
        return None

    hospital_data = {
        "三级甲等医院": copy.deepcopy(summary_data_field_data_type),
        "三级公立医院": copy.deepcopy(summary_data_field_data_type),
        "三级民营医院": copy.deepcopy(summary_data_field_data_type),
        "二级公立医院": copy.deepcopy(summary_data_field_data_type),
        "二级民营医院": copy.deepcopy(summary_data_field_data_type),
    }


    for row_dict in get_one_row_data(large_data):
        医疗机构名称 = row_dict["医疗机构名称"]
        医疗机构分类 = row_dict["医疗机构分类"]
        是否参加国家卫生健康委员会临床检验中心室间质量评价并合格 = row_dict["是否参加国家卫生健康委员会临床检验中心室间质量评价并合格"]
        通过国家室间质评医学检验结果合格项目数量 = row_dict["通过国家室间质评医学检验结果合格项目数量"]
        是否参加辽宁省临床检验中心室间质量评价并合格 = row_dict["是否参加辽宁省临床检验中心室间质量评价并合格"]
        通过辽宁省室间质评医学检验结果合格项目数量 = row_dict["通过辽宁省室间质评医学检验结果合格项目数量"]
        是否参加辽宁省医学影像质控中心影像质控认证评价并合格 = row_dict["是否参加辽宁省医学影像质控中心影像质控认证评价并合格"]
        要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可 = row_dict["要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可"]
        DR互认项目数 = row_dict["DR互认项目数"]
        DR节约检查费用 = row_dict["DR节约检查费用"]
        MR互认项目数 = row_dict["MR互认项目数"]
        MR节约检查费用 = row_dict["MR节约检查费用"]
        CT互认项目数 = row_dict["CT互认项目数"]
        CT节约检查费用 = row_dict["CT节约检查费用"]
        临床检验互认项目数 = row_dict["临床检验互认项目数"]
        临床检验节约检查费用 = row_dict["临床检验节约检查费用"]
        通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频 = row_dict["通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频"]
        总项目 = row_dict["总项目"]
        总费用 = row_dict["总费用"]
        Process_data = (
            f"\n <---- .... ---->\n"
            f"医疗机构名称: {医疗机构名称}\n"
            f"    医疗机构分类: {医疗机构分类}\n"
            f"    是否参加国家卫生健康委员会临床检验中心室间质量评价并合格: {是否参加国家卫生健康委员会临床检验中心室间质量评价并合格}\n"
            f"    通过国家室间质评医学检验结果合格项目数量: {通过国家室间质评医学检验结果合格项目数量}\n"
            f"    是否参加辽宁省临床检验中心室间质量评价并合格: {是否参加辽宁省临床检验中心室间质量评价并合格}\n"
            f"    通过辽宁省室间质评医学检验结果合格项目数量: {通过辽宁省室间质评医学检验结果合格项目数量}\n"
            f"    是否参加辽宁省医学影像质控中心影像质控认证评价并合格: {是否参加辽宁省医学影像质控中心影像质控认证评价并合格}\n"
            f"    要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可: {要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可}\n"
            f"    DR互认项目数: {DR互认项目数}\n"
            f"    DR节约检查费用: {DR节约检查费用}\n"
            f"    MR互认项目数: {MR互认项目数}\n"
            f"    MR节约检查费用: {MR节约检查费用}\n"
            f"    CT互认项目数: {CT互认项目数}\n"
            f"    CT节约检查费用: {CT节约检查费用}\n"
            f"    临床检验互认项目数: {临床检验互认项目数}\n"
            f"    临床检验节约检查费用: {临床检验节约检查费用}\n"
            f"    通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频: {通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频}\n"
            f"    总项目: {总项目}\n"
            f"    总费用: {总费用}\n"
        )
        processing_logger.info(Process_data)

        hospital_data[医疗机构分类]["医院总数量"] += 1
        if 是否参加辽宁省医学影像质控中心影像质控认证评价并合格 not in [0, 1]:
            raise ValueError(f"Invalid value for '是否参加辽宁省医学影像质控中心影像质控认证评价并合格': {是否参加辽宁省医学影像质控中心影像质控认证评价并合格}")

        if 是否参加国家卫生健康委员会临床检验中心室间质量评价并合格:
            hospital_data[医疗机构分类]["参加国家卫生健康委员会临床检验中心室间质量评价合格医院数量"] += 1
        hospital_data[医疗机构分类]["通过国家室间质评平均合格项目数量"] += int(通过国家室间质评医学检验结果合格项目数量)
        hospital_data[医疗机构分类]["参加辽宁省临床检验中心室间质量评价合格医院数量"] += int(是否参加辽宁省临床检验中心室间质量评价并合格)
        hospital_data[医疗机构分类]["通过辽宁省室间质评平均合格项目数量"] += int(通过辽宁省室间质评医学检验结果合格项目数量)
        hospital_data[医疗机构分类]["参加辽宁省医学影像质控中心影像质控认证评价合格医院数量"] += int(是否参加辽宁省医学影像质控中心影像质控认证评价并合格)
        hospital_data[医疗机构分类]["要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可的医院数量"] += int(要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可)
        hospital_data[医疗机构分类]["认可其他医疗机构标有互认标识的医学影像检查资料和医学检验结果"] += int(DR互认项目数) + int(MR互认项目数) + int(CT互认项目数) + int(临床检验互认项目数)
        hospital_data[医疗机构分类]["联盟医院间通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频次合计"] += int(通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频)
        hospital_data[医疗机构分类]["实施检查检验结果互认为患者节约医疗费用"] += float(DR节约检查费用) + float(MR节约检查费用) + float(CT节约检查费用) + float(临床检验节约检查费用)

        warning_process_data = (
            f"\n <---- .... ---->\n"
            f"医疗机构分类: {医疗机构分类}\n"
            f"    医院总数量: {hospital_data[医疗机构分类]['医院总数量']}\n"
            f"    参加国家卫生健康委员会临床检验中心室间质量评价合格医院数量: {hospital_data[医疗机构分类]['参加国家卫生健康委员会临床检验中心室间质量评价合格医院数量']}\n"
            f"    通过国家室间质评平均合格项目数量: {hospital_data[医疗机构分类]['通过国家室间质评平均合格项目数量']}\n"
            f"    参加辽宁省临床检验中心室间质量评价合格医院数量: {hospital_data[医疗机构分类]['参加辽宁省临床检验中心室间质量评价合格医院数量']}\n"
            f"    通过辽宁省室间质评平均合格项目数量: {hospital_data[医疗机构分类]['通过辽宁省室间质评平均合格项目数量']}\n"
            f"    参加辽宁省医学影像质控中心影像质控认证评价合格医院数量: {hospital_data[医疗机构分类]['参加辽宁省医学影像质控中心影像质控认证评价合格医院数量']}\n"
            f"    要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可的医院数量: {hospital_data[医疗机构分类]['要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可的医院数量']}\n"
            f"    认可其他医疗机构标有互认标识的医学影像检查资料和医学检验结果: {hospital_data[医疗机构分类]['认可其他医疗机构标有互认标识的医学影像检查资料和医学检验结果']}\n"
            f"    联盟医院间通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频次合计: {hospital_data[医疗机构分类]['联盟医院间通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频次合计']}\n"
            f"    实施检查检验结果互认为患者节约医疗费用: {hospital_data[医疗机构分类]['实施检查检验结果互认为患者节约医疗费用']}\n"   
        )
        # processing_logger.warning(warning_process_data)
    summary_data = dict()
    for key, value in hospital_data.items():
        value_list = []
        value_list.append(value["医院总数量"])
        value_list.append(value["参加国家卫生健康委员会临床检验中心室间质量评价合格医院数量"])
        value_list.append(value["通过国家室间质评平均合格项目数量"])
        value_list.append(value["参加辽宁省临床检验中心室间质量评价合格医院数量"])
        value_list.append(value["通过辽宁省室间质评平均合格项目数量"])
        value_list.append(value["参加辽宁省医学影像质控中心影像质控认证评价合格医院数量"])
        value_list.append(value["要求对其他医疗机构标有互认标识的医学影像检查资料和医学检验结果认可的医院数量"])
        value_list.append(value["认可其他医疗机构标有互认标识的医学影像检查资料和医学检验结果"])
        value_list.append(value["联盟医院间通过信息系统调阅成员医院间医学影像检查资料和医学检验结果频次合计"])
        value_list.append(value["实施检查检验结果互认为患者节约医疗费用"])
        summary_data[key] = value_list

    return summary_data


@app.route('/')
def index():
    response = make_response(render_template('index.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

def stream_logs(session_id):
    while True:
        if session_id not in log_streams:
            break
        log_stream = log_streams[session_id]
        logs = log_stream.read()
        if logs:
            for line in logs.splitlines():
                yield f"data: {line}\n\n"
            # Clearing the log stream for the current session while retaining the log data
        time.sleep(1)

@app.route('/logs')
def logs():
    session_id = get_session_id()
    response = Response(stream_logs(session_id), mimetype='text/event-stream')
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

def get_xlwt_style(xlrd_book, xlrd_sheet, row, col):
    """Get the xlwt style from an xlrd cell."""
    # Get the XF index of the cell in xlrd
    xf_index = xlrd_sheet.cell_xf_index(row, col)
    xf = xlrd_book.xf_list[xf_index]
    
    # Create an xlwt font
    font = xlrd_book.font_list[xf.font_index]
    xlwt_font = xlwt.Font()
    xlwt_font.height = font.height
    xlwt_font.italic = font.italic
    xlwt_font.struck_out = font.struck_out
    xlwt_font.outline = font.outline
    xlwt_font.shadow = font.shadow
    xlwt_font.colour_index = font.colour_index
    xlwt_font.bold = font.bold
    xlwt_font._weight = font.weight
    xlwt_font.escapement = font.escapement
    xlwt_font.underline = font.underline_type
    xlwt_font.family = font.family  # Preserve font family
    xlwt_font.charset = font.character_set

    # Create an xlwt borders
    borders = xlwt.Borders()
    borders.left = xf.border.left_line_style
    borders.right = xf.border.right_line_style
    borders.top = xf.border.top_line_style
    borders.bottom = xf.border.bottom_line_style
    borders.left_colour = xf.border.left_colour_index
    borders.right_colour = xf.border.right_colour_index
    borders.top_colour = xf.border.top_colour_index
    borders.bottom_colour = xf.border.bottom_colour_index

    # Create an xlwt alignment
    alignment = xlwt.Alignment()
    alignment.horz = xf.alignment.hor_align
    alignment.vert = xf.alignment.vert_align
    alignment.wrap = xf.alignment.text_wrapped

    # Create an xlwt pattern
    pattern = xlwt.Pattern()
    pattern.pattern = xf.background.fill_pattern
    pattern.pattern_fore_colour = xf.background.pattern_colour_index
    pattern.pattern_back_colour = xf.background.background_colour_index

    # Create an xlwt style with the font, borders, alignment, and pattern
    style = xlwt.XFStyle()
    style.font = xlwt_font
    style.borders = borders
    style.alignment = alignment
    style.pattern = pattern

    return style

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        session_id = get_session_id()
        stop_event = get_stop_event(session_id)
        processing_logger = get_logger(session_id)
        stop_event.clear()
        log_streams[session_id].truncate(0)
        log_streams[session_id].seek(0)

        large_excel = request.files.get('large_excel')
        summary_excel = request.files.get('summary_excel')

        if not large_excel or not summary_excel:
            processing_logger.error("Both files are required.")
            return jsonify({'error': 'Both files are required.'}), 400

        if not allowed_file(large_excel.filename) or not allowed_file(summary_excel.filename):
            processing_logger.error("Both files must be in Excel format (.xls or .xlsx).")
            return jsonify({'error': 'Both files must be in Excel format (.xls or .xlsx).'}), 400

        large_excel_path = os.path.join(UPLOAD_FOLDER, large_excel.filename)
        summary_excel_path = os.path.join(UPLOAD_FOLDER, summary_excel.filename)
        large_excel.save(large_excel_path)
        summary_excel.save(summary_excel_path)

        # Process each file individually
        large_data, large_book, large_sheet = read_excel(large_excel_path, processing_logger, stop_event)
        if large_data is None:
            return jsonify({'error': 'Error reading large Excel file or processing stopped.'}), 400
        validate_large_data(large_excel.filename, large_data, processing_logger)

        summary_data, summary_book, summary_sheet = read_excel(summary_excel_path, processing_logger, stop_event)
        if summary_data is None:
            return jsonify({'error': 'Error reading summary Excel file or processing stopped.'}), 400
        validate_summary_data(summary_excel.filename, summary_data, processing_logger)

        # Summarize large Excel data
        summarized_data = summarize_large_data(large_data, processing_logger, stop_event)
        if summarized_data is None:
            return jsonify({'error': 'Error summarizing large Excel data or processing stopped.'}), 400

        # Write the summarized data to the summary Excel file, preserving the format
        output_path = summary_excel_path  # Preserve the same file name and path

        if summary_excel.filename.endswith('.xlsx'):
            summary_sheet = summary_book.active
            # If the first column data value of summary_sheet is in summarized_data, use summarized_data instead
            for row in summary_sheet.iter_rows():
                if row[0].value in summarized_data:
                    key = row[0].value
                    row_data = summarized_data[key]
                    for col_idx, value in enumerate(row_data):
                        row[col_idx].value = value
            summary_book.save(output_path)
        else:
            wb = xl_copy(summary_book)
            ws = wb.get_sheet(0)
            # write 三级甲等医院 data to summary sheet
            for idx, value in enumerate(summarized_data["三级甲等医院"], 1):
                style = get_xlwt_style(summary_book, summary_sheet, 4, 0)
                ws.write(4, idx, value, style)
            # write 三级公立医院 data to summary sheet
            for idx, value in enumerate(summarized_data["三级公立医院"], 1):
                style = get_xlwt_style(summary_book, summary_sheet, 5, 0)
                ws.write(5, idx, value, style)
            for idx, value in enumerate(summarized_data["三级民营医院"], 1):
                style = get_xlwt_style(summary_book, summary_sheet, 6, 0)
                ws.write(6, idx, value, style)
            # write 二级公立医院 data to summary sheet
            for idx, value in enumerate(summarized_data["二级公立医院"], 1):
                style = get_xlwt_style(summary_book, summary_sheet, 7, 0)
                ws.write(7, idx, value, style)
            # write 二级民营医院 data to summary sheet
            for idx, value in enumerate(summarized_data["二级民营医院"], 1):
                style = get_xlwt_style(summary_book, summary_sheet, 8, 0)
                ws.write(8, idx, value, style)
            wb.save(output_path)

        # Clean up uploaded files
        os.remove(large_excel_path)
        response = send_file(output_path, as_attachment=True, attachment_filename=summary_excel.filename)
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        time.sleep(5)
        return response
    except Exception as e:
        session_id = get_session_id()
        processing_logger = get_logger(session_id)
        error_message = str(e)
        stack_trace = traceback.format_exc()
        processing_logger.error(f"Error: {error_message}")
        processing_logger.error(f"Stack trace:\n{stack_trace}")
        return jsonify({'error': error_message, 'stack_trace': stack_trace}), 500

@app.route('/stop', methods=['POST'])
def stop_processing():
    session_id = get_session_id()
    stop_event = get_stop_event(session_id)
    processing_logger = get_logger(session_id)
    stop_event.set()
    processing_logger.info("Processing stopped by user.")
    response = jsonify({'message': 'Processing stopped'})
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/clear_logs', methods=['POST'])
def clear_logs():
    session_id = get_session_id()
    if session_id in log_streams:
        log_streams[session_id].truncate(0)
        log_streams[session_id].seek(0)
    return jsonify({'message': 'Logs cleared'})

if __name__ == '__main__':
    # Disable Flask's default logging
    # log = logging.getLogger('werkzeug')
    # log.setLevel(logging.ERROR)
    app.run(debug=True, host='0.0.0.0', port=8000)
